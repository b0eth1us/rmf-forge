from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.finding import Finding, FindingStatus
from app.services.cci_mapper import map_cwe_to_ccis
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, uuid, zipfile
from datetime import datetime
from collections import Counter

router = APIRouter()

SEV_FILLS = {
    "critical":      PatternFill("solid", fgColor="C00000"),
    "high":          PatternFill("solid", fgColor="FF0000"),
    "medium":        PatternFill("solid", fgColor="FFC000"),
    "low":           PatternFill("solid", fgColor="FFFF00"),
    "informational": PatternFill("solid", fgColor="D9D9D9"),
}
HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
HDR_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="AAAAAA")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _set_header(ws, headers: list[str], widths: list[int]):
    ws.append(headers)
    for i, (cell, w) in enumerate(zip(ws[1], widths), 1):
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _style_row(ws, row_num: int, severity: str, n_cols: int):
    sev_key = (severity or "").lower()
    fill = SEV_FILLS.get(sev_key)
    for i in range(1, n_cols + 1):
        cell = ws.cell(row=row_num, column=i)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if fill and i == 3:  # only color the severity column
            cell.fill = fill


def _build_excel(findings: list[Finding], project_name: str) -> bytes:
    wb = Workbook()

    # ── Combined sheet ──────────────────────────────────────
    ws = wb.active
    ws.title = "All Findings"
    headers = ["Tool","ID","Severity","Title","File / URL","Line",
               "CWE","CVE","CCI","NIST Control","STIG Vuln ID",
               "Audit Action","Status","Justification","Comment","Code Snippet","Taint Trace"]
    widths =  [10,    8,   10,      40,     40,          8,
               8,     14,  12,      14,      14,
               14,          16,      40,           40,      30,            40]
    _set_header(ws, headers, widths)

    for r, f in enumerate(findings, 2):
        row = [
            f.source_tool, str(f.id)[:8], f.severity, f.title,
            f.file_path or f.affected_url or "",
            f.line_number or "",
            f.cwe_id, f.cve_id, f.cci_id, f.nist_control, f.vuln_id,
            f.audit_action, f.status.value, f.justification,
            f.audit_comment, f.code_snippet, f.taint_trace,
        ]
        ws.append([v or "" for v in row])
        _style_row(ws, r, f.severity, len(headers))

    # ── Per-tool sheets ──────────────────────────────────────
    for tool_name in ("fortify", "zap", "dep_check"):
        tool_findings = [f for f in findings if f.source_tool == tool_name]
        if not tool_findings:
            continue
        tws = wb.create_sheet(tool_name.replace("_", " ").title())
        if tool_name == "fortify":
            th = ["Severity","Title","File","Line","Taint Trace","Code Snippet","Audit Action","Developer Comment","Status","Justification"]
            tw = [10,40,40,8,50,40,14,50,16,40]
            _set_header(tws, th, tw)
            for r, f in enumerate(tool_findings, 2):
                tws.append([f.severity,f.title,f.file_path,f.line_number,
                            f.taint_trace,f.code_snippet,f.audit_action,
                            f.audit_comment,f.status.value,f.justification or ""])
                _style_row(tws, r, f.severity, len(th))
        elif tool_name == "zap":
            th = ["Severity","Alert","URL","CWE","CCI","NIST Control","Status","Justification"]
            tw = [10,40,50,8,12,14,16,40]
            _set_header(tws, th, tw)
            for r, f in enumerate(tool_findings, 2):
                tws.append([f.severity,f.title,f.affected_url,f.cwe_id,
                            f.cci_id,f.nist_control,f.status.value,f.justification or ""])
                _style_row(tws, r, f.severity, len(th))
        else:  # dep_check
            th = ["Severity","CVE","Dependency","Version","CWE","CCI","NIST Control","Description","Status","Justification"]
            tw = [10,18,30,12,8,12,14,50,16,40]
            _set_header(tws, th, tw)
            for r, f in enumerate(tool_findings, 2):
                tws.append([f.severity,f.cve_id,f.dependency_name,f.dependency_version,
                            f.cwe_id,f.cci_id,f.nist_control,f.description,
                            f.status.value,f.justification or ""])
                _style_row(tws, r, f.severity, len(th))

    # ── POA&M sheet ──────────────────────────────────────────
    open_findings = [f for f in findings if f.status == FindingStatus.open]
    if open_findings:
        pws = wb.create_sheet("POA&M")
        ph = ["ID","Tool","Severity","Weakness","CCI","NIST Control","STIG Vuln ID",
              "Scheduled Completion","Milestone Description","Justification"]
        pw = [8,10,10,50,12,14,14,18,50,40]
        _set_header(pws, ph, pw)
        for r, f in enumerate(open_findings, 2):
            scd = f.scheduled_completion_date.isoformat() if f.scheduled_completion_date else ""
            pws.append([str(f.id)[:8],f.source_tool,f.severity,f.title,
                        f.cci_id,f.nist_control,f.vuln_id,scd,
                        f.milestone_description or "",f.justification or ""])
            _style_row(pws, r, f.severity, len(ph))

    # ── Summary sheet ────────────────────────────────────────
    sws = wb.create_sheet("Summary", 0)
    sws.column_dimensions["A"].width = 28
    sws.column_dimensions["B"].width = 14
    sws.append(["RMF Forge — Evidence Package"])
    sws["A1"].font = Font(bold=True, size=14)
    sws.append(["Project", project_name])
    sws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")])
    sws.append(["Total Findings", len(findings)])
    sws.append([])
    sws.append(["By Status", ""])
    for k, v in Counter(f.status.value for f in findings).items():
        sws.append([k, v])
    sws.append([])
    sws.append(["By Severity", ""])
    for k, v in Counter((f.severity or "Unknown") for f in findings).most_common():
        sws.append([k, v])
    sws.append([])
    sws.append(["By Tool", ""])
    for k, v in Counter(f.source_tool for f in findings).items():
        sws.append([k, v])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_pdf(findings: list[Finding], project_name: str) -> bytes:
    """
    Build a PDF summary report using ReportLab.
    Falls back to a plain-text PDF if ReportLab is unavailable.
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=letter,
                                leftMargin=0.75*inch, rightMargin=0.75*inch,
                                topMargin=0.75*inch, bottomMargin=0.75*inch)
        styles = getSampleStyleSheet()
        story = []

        # Cover
        title_style = ParagraphStyle("Title", parent=styles["Title"],
                                     fontSize=22, spaceAfter=12, alignment=TA_CENTER)
        sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                   fontSize=12, alignment=TA_CENTER, textColor=colors.grey)
        story.append(Spacer(1, 1.5*inch))
        story.append(Paragraph("RMF Forge", title_style))
        story.append(Paragraph("Evidence Package", title_style))
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph(f"Project: {project_name}", sub_style))
        story.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", sub_style))
        story.append(PageBreak())

        h1 = styles["Heading1"]
        h2 = styles["Heading2"]
        normal = styles["Normal"]

        # Executive Summary
        story.append(Paragraph("Executive Summary", h1))
        story.append(Spacer(1, 0.1*inch))

        status_counts = Counter(f.status.value for f in findings)
        sev_counts = Counter((f.severity or "Unknown") for f in findings)
        tool_counts = Counter(f.source_tool for f in findings)

        summary_data = [["Metric", "Value"]]
        summary_data.append(["Total Findings", str(len(findings))])
        for k, v in status_counts.items():
            summary_data.append([f"  {k}", str(v)])
        summary_data.append(["", ""])
        for k, v in sev_counts.most_common():
            summary_data.append([f"  {k}", str(v)])

        t = Table(summary_data, colWidths=[3.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F3864")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8F8F8")]),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.2*inch))

        # Open findings by tool
        open_findings = [f for f in findings if f.status == FindingStatus.open]
        if open_findings:
            story.append(Paragraph("Open Findings", h1))
            for tool in ("fortify", "zap", "dep_check"):
                tool_open = [f for f in open_findings if f.source_tool == tool]
                if not tool_open:
                    continue
                story.append(Paragraph(tool.replace("_", " ").title(), h2))
                for f in tool_open:
                    story.append(Paragraph(
                        f"<b>[{f.severity or '?'}]</b> {f.title or 'Untitled'}", normal))
                    if f.file_path:
                        story.append(Paragraph(
                            f"&nbsp;&nbsp;File: {f.file_path}" +
                            (f" Line: {f.line_number}" if f.line_number else ""), normal))
                    if f.taint_trace:
                        story.append(Paragraph(f"&nbsp;&nbsp;Trace: {f.taint_trace}", normal))
                    if f.audit_comment:
                        story.append(Paragraph(f"&nbsp;&nbsp;Dev comment: {f.audit_comment}", normal))
                    if f.justification:
                        story.append(Paragraph(f"&nbsp;&nbsp;Justification: {f.justification}", normal))
                    story.append(Spacer(1, 0.05*inch))

        # Full findings table
        story.append(PageBreak())
        story.append(Paragraph("All Findings", h1))
        tbl_data = [["Sev","Tool","Title","CCI","Status"]]
        for f in findings:
            tbl_data.append([
                f.severity or "", f.source_tool,
                (f.title or "")[:60],
                f.cci_id or "", f.status.value,
            ])

        ft = Table(tbl_data, colWidths=[0.7*inch, 0.8*inch, 3.2*inch, 1.1*inch, 1.2*inch])
        sev_row_styles = []
        sev_color_map = {
            "Critical": "#C00000", "High": "#FF6B6B",
            "Medium": "#FFC000", "Low": "#FFFFAA",
        }
        for i, f in enumerate(findings, 1):
            hex_color = sev_color_map.get(f.severity or "", None)
            if hex_color:
                sev_row_styles.append(
                    ("BACKGROUND", (0, i), (0, i), colors.HexColor(hex_color))
                )
        ft.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F3864")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("GRID", (0,0), (-1,-1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8F8F8")]),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("WORDWRAP", (2,0), (2,-1), True),
            *sev_row_styles,
        ]))
        story.append(ft)

        doc.build(story)
        buf.seek(0)
        return buf.read()

    except ImportError:
        # ReportLab not installed — return minimal text-based PDF placeholder
        content = f"""%PDF-1.4
1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj
2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj
3 0 obj<</Type/Page/MediaBox[0 0 612 792]/Parent 2 0 R/Resources<</Font<</F1<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>>>>>>/Contents 4 0 R>>endobj
4 0 obj<</Length 200>>
stream
BT /F1 16 Tf 72 720 Td (RMF Forge — Evidence Package) Tj
0 -30 Td /F1 12 Tf (Project: {project_name}) Tj
0 -20 Td (Total Findings: {len(findings)}) Tj
0 -20 Td (Generated: {datetime.utcnow().strftime('%Y-%m-%d')}) Tj
0 -20 Td (Install reportlab for full PDF: pip install reportlab) Tj
ET
endstream
endobj
xref
0 5
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000274 00000 n
trailer<</Size 5/Root 1 0 R>>
startxref
525
%%EOF"""
        return content.encode()


@router.post("/package/{project_id}")
def export_package(project_id: uuid.UUID, db: Session = Depends(get_db)):
    """Generate the full evidence package: Excel + PDF + .ckl bundled in a ZIP."""
    from app.models.project import Project
    from app.api.routes.stig import _build_ckl

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    findings = db.query(Finding).filter(Finding.project_id == project_id)\
        .order_by(Finding.severity, Finding.source_tool).all()

    safe_name = project.name.replace(" ", "_")

    excel_bytes = _build_excel(findings, project.name)
    pdf_bytes   = _build_pdf(findings, project.name)
    ckl_bytes   = _build_ckl(findings, project)

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{safe_name}_findings.xlsx", excel_bytes)
        zf.writestr(f"{safe_name}_report.pdf",    pdf_bytes)
        zf.writestr(f"{safe_name}.ckl",           ckl_bytes)

    zip_buf.seek(0)
    return Response(
        content=zip_buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_evidence_package.zip"'},
    )


@router.get("/consolidated/{project_id}")
def export_consolidated(project_id: uuid.UUID, db: Session = Depends(get_db)):
    """Quick Excel-only export (no PDF, no ZIP)."""
    from app.models.project import Project
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")
    findings = db.query(Finding).filter(Finding.project_id == project_id).all()
    excel_bytes = _build_excel(findings, project.name)
    safe_name = project.name.replace(" ", "_")
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_findings.xlsx"'},
    )


@router.get("/emass-zap/{project_id}")
def export_emass_zap(project_id: uuid.UUID, db: Session = Depends(get_db)):
    """eMASS bulk upload template for ZAP findings."""
    from app.models.project import Project
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(404, "Project not found")

    findings = db.query(Finding).filter(
        Finding.project_id == project_id,
        Finding.source_tool.in_(["zap", "zap_xml", "zap_json"]),
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "eMASS Import"
    headers = ["Control Number","CCI","Implementation Status",
               "Implementation Narrative","Finding","Risk Description",
               "Mitigation","Severity","Relevance of Threat"]
    widths =  [16,           12,   20,
               50,                      50,          50,
               30,          10,       20]
    _set_header(ws, headers, widths)

    for f in findings:
        ccis = map_cwe_to_ccis(f.cwe_id or "")
        for cci_info in (ccis or [{}]):
            ws.append([
                cci_info.get("nist_control", ""),
                cci_info.get("cci_id", ""),
                f.status.value,
                f.justification or "",
                f.title or "",
                f.description or "",
                "",
                f.severity or "",
                "",
            ])

    ws2 = wb.create_sheet("Unmapped CWEs")
    _set_header(ws2, ["Title","CWE ID","Severity","Notes"], [40,10,10,50])
    for f in findings:
        if not map_cwe_to_ccis(f.cwe_id or ""):
            ws2.append([f.title, f.cwe_id, f.severity, "No CCI mapping — review manually"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = project.name.replace(" ", "_")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_emass.xlsx"'},
    )
