from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, Form
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.project import Project
from app.models.finding import Finding, FindingStatus
from app.models.import_log import ImportLog
from app.services.column_mapper import suggest_mapping
from app.services.finding_hasher import stable_key
from app.services.stig_mapper import map_finding_to_stig
from app.services.cci_mapper import map_cwe_to_ccis
from app.parsers.fortify_parser import parse_fpr
from app.parsers.zap_parser import parse_zap_xml, parse_zap_json
from app.parsers.dep_check_parser import parse_dep_check_xml, parse_dep_check_json
import hashlib, json, io, uuid
from datetime import datetime
from openpyxl import load_workbook
import csv as csvlib

router = APIRouter()

def _hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def _parse_csv(data: bytes) -> list[dict]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csvlib.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]

def _parse_xlsx(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]

def _detect_tool(filename: str, data: bytes) -> str:
    fn = filename.lower()
    if fn.endswith(".fpr"):
        return "fortify"
    if fn.endswith(".json"):
        # Peek content to distinguish ZAP vs dep-check JSON
        try:
            obj = json.loads(data[:4096])
            if "dependencies" in obj:
                return "dep_check_json"
            if "site" in obj or "alerts" in obj:
                return "zap_json"
        except Exception:
            pass
        return "zap_json"
    if fn.endswith(".xml"):
        if b"dependency-check" in data[:1024] or b"DependencyCheckReport" in data[:1024]:
            return "dep_check_xml"
        if b"OWASPZAPReport" in data[:512] or b"alertitem" in data[:512]:
            return "zap_xml"
        return "xml_generic"
    if fn.endswith(".csv"):
        return "csv"
    if fn.endswith((".xlsx", ".xls")):
        return "xlsx"
    return "unknown"

def _parse_file(tool: str, data: bytes) -> list[dict]:
    if tool == "fortify":       return parse_fpr(data)
    if tool == "zap_xml":       return parse_zap_xml(data)
    if tool == "zap_json":      return parse_zap_json(data)
    if tool == "dep_check_xml": return parse_dep_check_xml(data)
    if tool == "dep_check_json":return parse_dep_check_json(data)
    if tool == "csv":           return _parse_csv(data)
    if tool == "xlsx":          return _parse_xlsx(data)
    raise HTTPException(status_code=400, detail=f"Unsupported file type: {tool}")

def _normalize_finding(raw: dict, tool: str, column_map: dict[str, str]) -> dict:
    # Parsers already return normalized dicts — column_map only applies to csv/xlsx
    if tool in ("csv", "xlsx"):
        out: dict = {"source_tool": tool}
        reverse = {v: k for k, v in column_map.items()}
        for col, value in raw.items():
            canonical = reverse.get(col, col)
            out[canonical] = str(value) if value is not None else ""
        return out
    return {**raw, "source_tool": tool}

def _auto_map(norm: dict) -> dict:
    """Derive vuln_id, cci_id, nist_control from finding content."""
    vuln_id = norm.get("vuln_id") or None
    cci_id = norm.get("cci_id") or None
    nist_control = norm.get("nist_control") or None

    cwe = re.sub(r'[^0-9]', '', str(norm.get("cwe_id") or ""))
    if cwe:
        ccis = map_cwe_to_ccis(cwe)
        if ccis:
            if not cci_id:
                cci_id = ccis[0].get("cci_id")
            if not nist_control:
                nist_control = ccis[0].get("nist_control")

    if not vuln_id or not cci_id:
        match = map_finding_to_stig(norm.get("title", ""), norm.get("description", ""))
        if not vuln_id and match.get("vuln_ids"):
            vuln_id = match["vuln_ids"][0]
        if not cci_id and match.get("cci_ids"):
            cci_id = match["cci_ids"][0]

    return {"vuln_id": vuln_id, "cci_id": cci_id, "nist_control": nist_control}

import re

@router.post("/preview")
async def preview_columns(file: UploadFile = File(...)):
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")
    tool = _detect_tool(file.filename or "", data)
    try:
        rows = _parse_file(tool, data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {str(e)}")
    if not rows:
        raise HTTPException(status_code=422, detail="No data found in file")
    columns = list(rows[0].keys())
    suggestions = suggest_mapping(columns) if tool in ("csv", "xlsx") else {}
    return {
        "filename": file.filename,
        "tool": tool,
        "row_count": len(rows),
        "columns": columns,
        "suggestions": suggestions,
        "file_hash": _hash(data),
    }


@router.post("/import/{project_id}")
async def import_file(
    project_id: uuid.UUID,
    file: UploadFile = File(...),
    column_map: str = Form(default="{}"),
    db: Session = Depends(get_db),
):
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    file_hash = _hash(data)
    tool = _detect_tool(file.filename or "", data)

    try:
        col_map: dict[str, str] = json.loads(column_map)
    except json.JSONDecodeError:
        col_map = {}

    try:
        raw_rows = _parse_file(tool, data)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {str(e)}")

    added = updated = unchanged = 0

    for row in raw_rows:
        norm = _normalize_finding(row, tool, col_map)
        key = stable_key(tool, norm.get("plugin_id"), norm.get("title"))
        mappings = _auto_map(norm)

        existing = db.query(Finding).filter(
            Finding.project_id == project_id,
            Finding.stable_key == key,
        ).first()

        if existing:
            existing.severity = norm.get("severity") or existing.severity
            existing.description = norm.get("description") or existing.description
            existing.last_seen = datetime.utcnow()
            # Update Fortify-specific fields always (re-import = fresh scan)
            if norm.get("file_path"):    existing.file_path = norm["file_path"]
            if norm.get("line_number"):  existing.line_number = norm["line_number"]
            if norm.get("code_snippet"): existing.code_snippet = norm["code_snippet"]
            if norm.get("taint_trace"):  existing.taint_trace = norm["taint_trace"]
            # Preserve existing audit comment if developer updated it externally
            if norm.get("audit_comment") and not existing.audit_comment:
                existing.audit_comment = norm["audit_comment"]
            if norm.get("audit_action"):
                existing.audit_action = norm["audit_action"]
            # Only auto-map if not manually set
            if not existing.vuln_id and mappings["vuln_id"]:
                existing.vuln_id = mappings["vuln_id"]
            if not existing.cci_id and mappings["cci_id"]:
                existing.cci_id = mappings["cci_id"]
            if not existing.nist_control and mappings["nist_control"]:
                existing.nist_control = mappings["nist_control"]
            unchanged += 1 if existing.justification else 0
            updated += 1 if not existing.justification else 0
        else:
            finding = Finding(
                project_id=project_id,
                stable_key=key,
                source_tool=tool,
                severity=norm.get("severity"),
                title=norm.get("title"),
                description=norm.get("description"),
                plugin_id=norm.get("plugin_id"),
                cwe_id=norm.get("cwe_id"),
                cve_id=norm.get("cve_id"),
                vuln_id=mappings["vuln_id"],
                cci_id=mappings["cci_id"],
                nist_control=mappings["nist_control"],
                # Fortify fields
                audit_comment=norm.get("audit_comment"),
                audit_action=norm.get("audit_action"),
                file_path=norm.get("file_path"),
                line_number=norm.get("line_number"),
                code_snippet=norm.get("code_snippet"),
                taint_trace=norm.get("taint_trace"),
                # ZAP / dep-check fields
                affected_url=norm.get("affected_url"),
                dependency_name=norm.get("dependency_name"),
                dependency_version=norm.get("dependency_version"),
                status=FindingStatus.not_reviewed,
                raw_data=json.dumps(row),
            )
            db.add(finding)
            added += 1

    log = ImportLog(
        project_id=project_id,
        filename=file.filename,
        file_hash=file_hash,
        source_tool=tool,
        findings_added=added,
        findings_updated=updated,
        findings_unchanged=unchanged,
    )
    db.add(log)
    db.commit()

    return {"status": "ok", "findings_added": added, "findings_updated": updated, "findings_unchanged": unchanged}
